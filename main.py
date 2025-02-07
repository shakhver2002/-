import re
import dns.resolver
import smtplib
import openpyxl
from openpyxl.styles import PatternFill

def validate_email(email):
    # Проверка формата email
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        return False, "Неверный формат email"

    # Извлечение домена
    domain = email.split('@')[1]

    try:
        # Проверка наличия MX-записей у домена
        mx_records = dns.resolver.resolve(domain, 'MX')
        mx_hosts = [str(record.exchange) for record in mx_records]
    except dns.resolver.NXDOMAIN:
        return False, "Домен не существует"
    except dns.resolver.NoAnswer:
        return False, "У домена отсутствуют MX-записи"
    except dns.exception.Timeout:
        return False, "Превышено время ожидания ответа от DNS"
    except dns.resolver.NoNameservers:
        return False, "Нет доступных серверов имен для проверки"
    except Exception as e:
        return False, f"Ошибка проверки: {str(e)}"

    # Проверка существования email через SMTP
    try:
        # Подключение к первому доступному MX-серверу
        smtp = smtplib.SMTP(timeout=10)
        smtp.connect(mx_hosts[0])
        smtp.helo()  # Приветствие
        smtp.mail("test@example.com")  # Указываем отправителя
        code, message = smtp.rcpt(email)  # Проверка получателя
        smtp.quit()

        # Код 250 указывает, что адрес существует
        if code == 250:
            return True, "Email существует и доступен для доставки"
        elif code == 550:
            return False, "Email не существует (ошибка 550: отказ сервера)"
        else:
            return False, f"Email недоступен (SMTP ответ: {code})"

    except smtplib.SMTPConnectError:
        return False, "Не удалось подключиться к серверу"
    except smtplib.SMTPServerDisconnected:
        return False, "Сервер неожиданно завершил соединение"
    except smtplib.SMTPException as e:
        return False, f"Ошибка SMTP: {str(e)}"
    except Exception as e:
        return False, f"Произошла непредвиденная ошибка: {str(e)}"

##Неинтересная настройка нового excel файла (делаем id, email, сообщение(Успех=зеленый, Ошибка=Красный, Желтый=Возможно email валидный, но сервер не может ответить))
def process_excel(input_file, output_file):
    # Открываем входной Excel файл
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Создаем новый Excel файл для результатов
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(["ID", "Email", "Сообщение"])

    # Цвета для успеха и ошибок
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    count = 1 # счетчик

    # Обработка строк файла
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[0]

        if email:
            is_valid, message = validate_email(email)
            new_row = [count, email, message]

            # Добавляем строку в новый файл
            new_ws.append(new_row)

            count += 1
            print(count, ' ', email) ## вывод счетчика, чтобы понимать сколько ждать

            # Окрашиваем строку в зависимости от результата
            if "Превышено время ожидания ответа от DNS" in message:
                fill = warning_fill  # желтый цвет для ошибки DNS таймаута
            elif is_valid:
                fill = success_fill  # зеленый цвет для успеха
            else:
                fill = error_fill  # красный цвет для ошибки

            for col_idx in range(1, 4):
                new_ws.cell(row=new_ws.max_row, column=col_idx).fill = fill

    # Сохраняем новый файл
    new_wb.save(output_file)

# Пример использования
input_file = "{your_file}"  # Исходный файл в формате xlsx
output_file = "validated_emails.xlsx"  # Итоговый файл
process_excel(input_file, output_file)

